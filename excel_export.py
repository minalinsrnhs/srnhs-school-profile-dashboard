"""Excel upload and export helpers for the SRNHS dashboard runtime.

The live online dashboard stores records centrally; this module creates a clean,
formula-driven Excel export and accepts the same workbook structure for updates.
"""
from __future__ import annotations

from io import BytesIO
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from analytics import compute_dashboard

GREEN = "168447"; DARK = "0B3D24"; MID = "58B878"; PALE = "F4FBF6"; LIGHT = "E5F4EA"; YELLOW = "FFF4CC"; WHITE = "FFFFFF"; LINE = "DCE8E0"


def _title(ws, title: str, subtitle: str, end_col: int = 9) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=end_col)
    cell = ws.cell(1, 1, title); cell.fill = PatternFill("solid", fgColor=DARK); cell.font = Font(color=WHITE, bold=True, size=18); cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32; ws.row_dimensions[2].height = 32
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=end_col)
    cell = ws.cell(3, 1, subtitle); cell.fill = PatternFill("solid", fgColor=LIGHT); cell.font = Font(color=DARK, italic=True, size=10); cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[3].height = 30


def _header(ws, row: int, labels: list[str]) -> None:
    for col, label in enumerate(labels, 1):
        cell = ws.cell(row, col, label); cell.fill = PatternFill("solid", fgColor=GREEN); cell.font = Font(color=WHITE, bold=True, size=10); cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = Border(bottom=Side(style="thin", color=LINE))
    ws.row_dimensions[row].height = 32


def _format_table(ws, start_row: int, end_row: int, start_col: int, end_col: int, computed_cols: set[int] | None = None) -> None:
    if end_row < start_row:
        return
    computed_cols = computed_cols or set()
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col); cell.fill = PatternFill("solid", fgColor=LIGHT if col in computed_cols else WHITE); cell.border = Border(bottom=Side(style="thin", color=LINE)); cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col in computed_cols: cell.font = Font(color=DARK, bold=True)
        ws.row_dimensions[row].height = 27


def _finalize(wb: Workbook) -> BytesIO:
    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = sheet.freeze_panes or "A6"
    output = BytesIO(); wb.save(output); output.seek(0); return output


def export_workbook(db, settings: dict[str, Any]) -> BytesIO:
    records, resources, cohorts = db.list_records(), db.list_resources(), db.list_cohort()
    actions, accounts, activity = db.list_actions(), db.list_accounts(), db.list_activity(250)
    years = sorted({r["school_year"] for r in records})
    analytics = compute_dashboard(records, resources, cohorts, None, "All")
    wb = Workbook(); wb.remove(wb.active)

    ws = wb.create_sheet("START_HERE")
    _title(ws, settings.get("dashboard_title", "SRNHS School Profile Analysis Dashboard"), "Simple editable workbook exported from the live dashboard.", 8)
    _header(ws, 5, ["Step", "What to Do"])
    steps = [
        ("1", "Edit or review school values in DATA_ENTRY, RESOURCES, COHORT_TRANSITION, and ACTION_PLANS."),
        ("2", "SUMMARY and PLANNING_TOOLS contain formulas and automatically calculate after input changes."),
        ("3", "Upload a completed workbook through School Records to add or update school-year values."),
        ("4", "Account passwords are never exported as readable text; only masked display is shown."),
        ("5", "The live online dashboard remains the central source when multiple users work at the same time."),
    ]
    for row, pair in enumerate(steps, 6): ws.cell(row,1,pair[0]); ws.cell(row,2,pair[1])
    _format_table(ws, 6, 10, 1, 2); ws.column_dimensions["A"].width = 14; ws.column_dimensions["B"].width = 96

    ws = wb.create_sheet("SETTINGS")
    _title(ws, "Dashboard Settings", "Editable display details used for the dashboard presentation.", 3)
    _header(ws, 5, ["Setting", "Current Value", "Where It Appears"])
    setting_rows = [("Dashboard Title", settings.get("dashboard_title", ""), "Top bar and reports"),("School Name", settings.get("school_name", ""), "Login and top bar"),("Location", settings.get("location", ""), "Login and top bar"),("Dashboard Subtitle", settings.get("subtitle", ""), "Top bar"),("Logo URL", settings.get("logo_url", ""), "Login and sidebar"),("Login Background URL", settings.get("login_background_url", ""), "Login page"),("Main Green", settings.get("main_green", "#168447"), "Buttons and accents"),("Sidebar Color", settings.get("sidebar_color", "#071B12"), "Sidebar"),("Page Background", settings.get("page_background", "#F8FBF8"), "Workspace")]
    for row, values in enumerate(setting_rows, 6):
        for col, value in enumerate(values, 1): ws.cell(row,col,value)
    _format_table(ws, 6, 14, 1, 3); [setattr(ws.cell(row,2), 'fill', PatternFill("solid", fgColor=YELLOW)) for row in range(6,15)]
    for col, width in {"A":30,"B":76,"C":30}.items(): ws.column_dimensions[col].width = width

    ws = wb.create_sheet("DATA_ENTRY")
    _title(ws, "School Records Input", "Yellow cells are editable inputs; green cells calculate rates automatically.", 9)
    _header(ws, 5, ["School Year", "Level", "Grade Level", "Enrollment", "Dropouts", "Repeaters", "Teachers", "Dropout Rate", "Repeater Rate"])
    records_sorted = sorted(records, key=lambda r: (r["school_year"], int(r["grade_level"].split()[-1])))
    row = 6
    for item in records_sorted:
        values = [item["school_year"], item["level"], item["grade_level"], int(item["enrollment"]), int(item["dropouts"]), int(item["repeaters"]), int(item["teachers"])]
        for col, value in enumerate(values, 1): ws.cell(row,col,value)
        ws.cell(row,8,f'=IF(D{row}="","",IFERROR(E{row}/D{row},0))'); ws.cell(row,9,f'=IF(D{row}="","",IFERROR(F{row}/D{row},0))'); row += 1
    last_year = int(years[-1].split("-")[0]) if years else 2025; next_year = f"{last_year+1}-{last_year+2}"
    if next_year not in years:
        for number in range(7,13):
            ws.cell(row,1,next_year); ws.cell(row,2,"JHS" if number<=10 else "SHS"); ws.cell(row,3,f"Grade {number}"); ws.cell(row,8,f'=IF(D{row}="","",IFERROR(E{row}/D{row},0))'); ws.cell(row,9,f'=IF(D{row}="","",IFERROR(F{row}/D{row},0))'); row += 1
    _format_table(ws,6,row-1,1,9,{8,9})
    for r in range(6,row):
        if ws.cell(r,1).value == next_year and next_year not in years:
            for col in range(1,8): ws.cell(r,col).fill = PatternFill("solid", fgColor=YELLOW)
        ws.cell(r,8).number_format = "0.00%"; ws.cell(r,9).number_format = "0.00%"
    for col,width in enumerate([17,12,16,14,13,13,13,16,17],1): ws.column_dimensions[get_column_letter(col)].width = width
    ws.auto_filter.ref = f"A5:I{row-1}"; ws.freeze_panes = "A6"

    ws = wb.create_sheet("RESOURCES")
    _title(ws, "Classrooms and Room Sizes", "Formula-based classroom totals and students-per-classroom results.", 7)
    _header(ws, 5, ["School Year", "JHS Classrooms", "SHS Classrooms", "Total Classrooms", "Total Enrollment", "Students per Classroom", "Notes"])
    res_map = {r["school_year"]: r for r in resources}
    for i, year in enumerate(years, 6):
        item=res_map.get(year,{})
        ws.cell(i,1,year); ws.cell(i,2,int(item.get("jhs_classrooms",0) or 0)); ws.cell(i,3,int(item.get("shs_classrooms",0) or 0)); ws.cell(i,4,f'=B{i}+C{i}'); ws.cell(i,5,f'=SUMIFS(DATA_ENTRY!$D:$D,DATA_ENTRY!$A:$A,A{i})'); ws.cell(i,6,f'=IFERROR(E{i}/D{i},0)')
    _format_table(ws,6,5+len(years),1,7,{4,5,6})
    start=9+len(years); _header(ws,start,["Grade Level", "Room Length (m)", "Room Width (m)", "Room Area (sqm)"])
    for i,item in enumerate(db.list_room_sizes(), start+1):
        ws.cell(i,1,item["grade_level"]); ws.cell(i,2,item["room_length"]); ws.cell(i,3,item["room_width"]); ws.cell(i,4,f'=B{i}*C{i}')
    _format_table(ws,start+1,start+len(db.list_room_sizes()),1,4,{4})
    for col,width in {"A":19,"B":19,"C":19,"D":20,"E":20,"F":24,"G":36}.items(): ws.column_dimensions[col].width=width

    ws = wb.create_sheet("COHORT_TRANSITION")
    _title(ws, "Cohort and Transition Trends", "Formula-driven continuity indicators.", 12)
    _header(ws, 5, ["School Year", "Grade 7 Baseline Year", "Grade 7 Baseline", "Grade 12 Enrollment", "Cohort Survival Rate", "Notes"])
    c_map = {c["school_year"]: c for c in cohorts}
    for i, year in enumerate(years, 6):
        item=c_map.get(year,{})
        ws.cell(i,1,year); ws.cell(i,2,item.get("baseline_year","")); ws.cell(i,3,item.get("grade7_baseline","")); ws.cell(i,4,item.get("grade12_current","")); ws.cell(i,5,f'=IFERROR(D{i}/C{i},"")'); ws.cell(i,5).number_format="0.00%"
    _format_table(ws,6,5+len(years),1,6,{5})
    for col,label in enumerate(["Current School Year", "Previous Grade 10 Enrollment", "Current Grade 11 Enrollment", "Transition Rate", "Plain Meaning"],8):
        ws.cell(5,col,label); ws.cell(5,col).fill=PatternFill("solid",fgColor=GREEN); ws.cell(5,col).font=Font(color=WHITE,bold=True); ws.cell(5,col).alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for i, year in enumerate(years[1:],6):
        prior=years[years.index(year)-1]; prev_g10=next((r["enrollment"] for r in records if r["school_year"]==prior and r["grade_level"]=="Grade 10"),0); cur_g11=next((r["enrollment"] for r in records if r["school_year"]==year and r["grade_level"]=="Grade 11"),0)
        ws.cell(i,8,year); ws.cell(i,9,prev_g10); ws.cell(i,10,cur_g11); ws.cell(i,11,f'=IFERROR(J{i}/I{i},"")'); ws.cell(i,11).number_format="0.00%"; ws.cell(i,12,"Grade 11 compared with previous Grade 10")
    _format_table(ws,6,5+len(years),8,12,{11})
    for col,width in {"A":19,"B":24,"C":20,"D":21,"E":21,"F":26,"H":21,"I":27,"J":27,"K":18,"L":48}.items(): ws.column_dimensions[col].width=width

    ws = wb.create_sheet("SUMMARY")
    _title(ws, "Dashboard Summary", "Chart-ready formula values that update from input sheets.", 15)
    _header(ws,5,["School Year","Total Enrollment","JHS","SHS","Enrollment Change %","Dropouts","Dropout Rate","Repeaters","Repeater Rate","Teachers","Students per Teacher","Classrooms","Students per Classroom","Cohort Survival","Transition Rate"])
    for i,year in enumerate(years,6):
        ws.cell(i,1,year); formulas=[f'=SUMIFS(DATA_ENTRY!$D:$D,DATA_ENTRY!$A:$A,A{i})',f'=SUMIFS(DATA_ENTRY!$D:$D,DATA_ENTRY!$A:$A,A{i},DATA_ENTRY!$B:$B,"JHS")',f'=SUMIFS(DATA_ENTRY!$D:$D,DATA_ENTRY!$A:$A,A{i},DATA_ENTRY!$B:$B,"SHS")','' if i==6 else f'=IFERROR(B{i}/B{i-1}-1,0)',f'=SUMIFS(DATA_ENTRY!$E:$E,DATA_ENTRY!$A:$A,A{i})',f'=IFERROR(F{i}/B{i},0)',f'=SUMIFS(DATA_ENTRY!$F:$F,DATA_ENTRY!$A:$A,A{i})',f'=IFERROR(H{i}/B{i},0)',f'=SUMIFS(DATA_ENTRY!$G:$G,DATA_ENTRY!$A:$A,A{i})',f'=IFERROR(B{i}/J{i},0)',f'=SUMIFS(RESOURCES!$D:$D,RESOURCES!$A:$A,A{i})',f'=IFERROR(B{i}/L{i},0)',f'=IFERROR(SUMIFS(COHORT_TRANSITION!$E:$E,COHORT_TRANSITION!$A:$A,A{i}),"")',f'=IFERROR(SUMIFS(COHORT_TRANSITION!$K:$K,COHORT_TRANSITION!$H:$H,A{i}),"")']
        for col,formula in enumerate(formulas,2):
            if formula: ws.cell(i,col,formula)
    _format_table(ws,6,5+len(years),1,15,set(range(2,16)))
    for col in [5,7,9,14,15]:
        for r in range(6,6+len(years)): ws.cell(r,col).number_format="0.00%"
    for col,width in enumerate([17,19,13,13,20,13,16,14,17,13,21,16,23,19,18],1): ws.column_dimensions[get_column_letter(col)].width=width
    chart=LineChart(); chart.title="Enrollment Trend"; chart.y_axis.title="Students"; chart.x_axis.title="School Year"; chart.add_data(Reference(ws,min_col=2,min_row=5,max_row=5+len(years)),titles_from_data=True); chart.set_categories(Reference(ws,min_col=1,min_row=6,max_row=5+len(years))); chart.height=8; chart.width=18; ws.add_chart(chart,"Q5")

    ws = wb.create_sheet("PLANNING_TOOLS")
    _title(ws,"Forecast and What-If Planning","Yellow cells are assumptions; green cells calculate planning outputs.",10)
    _header(ws,5,["Index","School Year","Enrollment","Type"])
    for i,year in enumerate(years,6): ws.cell(i,1,i-5); ws.cell(i,2,year); ws.cell(i,3,f'=SUMMARY!B{i}'); ws.cell(i,4,"Actual")
    forecast_row=6+len(years); last_start=int(years[-1].split("-")[0])
    for offset in range(1,3):
        r=forecast_row+offset-1; ws.cell(r,1,len(years)+offset); ws.cell(r,2,f"{last_start+offset}-{last_start+offset+1}"); ws.cell(r,3,f'=FORECAST.LINEAR(A{r},C6:C{5+len(years)},A6:A{5+len(years)})'); ws.cell(r,4,"Projected")
    _format_table(ws,6,forecast_row+1,1,4,{3})
    _header(ws,5,["Index","School Year","Enrollment","Type"])
    ws.cell(5,6,"WHAT-IF TOOL"); ws.cell(5,6).fill=PatternFill("solid",fgColor=GREEN); ws.cell(5,6).font=Font(color=WHITE,bold=True)
    planning=[(7,"Base Enrollment",f'=SUMMARY!B{5+len(years)}',"Latest year total"),(8,"Assumed Change",-0.05,"Edit this rate"),(9,"Scenario Enrollment",'=G7*(1+G8)',"Computed result"),(11,"Target Repeater Rate",0.015,"Edit this rate"),(12,"Target Repeaters",'=ROUND(G9*G11,0)',"Computed result"),(14,"Assumed Teachers",f'=SUMMARY!J{5+len(years)}',"Edit this number"),(15,"Students per Teacher",'=IFERROR(G9/G14,0)',"Computed result")]
    for row,label,value,note in planning:
        ws.cell(row,6,label);ws.cell(row,7,value);ws.cell(row,8,note);_format_table(ws,row,row,6,8,{7})
    for row in [8,11,14]: ws.cell(row,7).fill=PatternFill("solid",fgColor=YELLOW)
    ws.cell(8,7).number_format="0.00%";ws.cell(11,7).number_format="0.00%"
    for col,width in {"A":12,"B":18,"C":18,"D":14,"F":29,"G":20,"H":30}.items(): ws.column_dimensions[col].width=width

    ws = wb.create_sheet("ACTION_PLANS")
    _title(ws,"Insights and Action Plans","Editable long-term action monitoring entries saved from the dashboard.",14)
    action_columns=["School Year","Analysis Type","Focus Area","Observed Pattern","Data Basis","Suggested Action","Responsible Group","Target Indicator","Baseline Value","Target Value","Monitoring Period","Current Result","Status","Progress Notes","Additional Notes"]
    _header(ws,5,action_columns)
    action_keys=["school_year","analysis_type","focus_area","observed_pattern","data_basis","suggested_action","responsible_group","target_indicator","baseline_value","target_value","monitoring_period","current_result","status","progress_notes","notes"]
    for row,item in enumerate(actions,6):
        for col,key in enumerate(action_keys,1): ws.cell(row,col,item.get(key,""))
    _format_table(ws,6,max(6,5+len(actions)),1,15)
    for col,width in enumerate([17,18,24,46,45,52,29,25,25,25,25,25,17,38,36],1): ws.column_dimensions[get_column_letter(col)].width=width

    ws = wb.create_sheet("INSIGHTS_SUMMARY")
    _title(ws,"Business Intelligence Summary","Dynamic analysis derived from currently saved records.",5)
    _header(ws,5,["Area","Analytics Type","Finding","Data Basis","Planning Use"])
    row=6
    for area,types in analytics["bi_tabs"].items():
        for kind,items in types.items():
            for item in items:
                ws.cell(row,1,area.title());ws.cell(row,2,kind.title());ws.cell(row,3,item["title"]+": "+item["text"]);ws.cell(row,4,item.get("basis",""));ws.cell(row,5,"Review in dashboard and action plans.");row+=1
    _format_table(ws,6,row-1,1,5)
    for col,width in {"A":20,"B":20,"C":95,"D":60,"E":38}.items(): ws.column_dimensions[col].width=width

    ws = wb.create_sheet("USERS")
    _title(ws,"Account Holders","Password display is masked. Actual passwords are never exported.",5)
    _header(ws,5,["Full Name","Username","Password Display","Photo Reference","Last Active"])
    for row,account in enumerate(accounts,6):
        values=[account["full_name"],account["username"],"••••••••", "Photo saved in dashboard" if account.get("avatar_url") else "", account.get("last_active","")]
        for col,value in enumerate(values,1): ws.cell(row,col,value)
    _format_table(ws,6,max(6,5+len(accounts)),1,5)
    for col,width in {"A":30,"B":22,"C":22,"D":30,"E":30}.items(): ws.column_dimensions[col].width=width

    ws = wb.create_sheet("ACTIVITY_LOG")
    _title(ws,"Recent Activity","Latest dashboard actions included in this export.",6)
    _header(ws,5,["Date and Time","User","Action","Section","School Year or Record","Details"])
    for row,item in enumerate(activity,6):
        for col,value in enumerate([item.get("occurred_at",""),item.get("display_name",""),item.get("action",""),item.get("section",""),item.get("affected_record",""),item.get("details","")],1): ws.cell(row,col,value)
    _format_table(ws,6,max(6,5+len(activity)),1,6)
    for col,width in {"A":27,"B":27,"C":32,"D":24,"E":26,"F":62}.items(): ws.column_dimensions[col].width=width
    return _finalize(wb)


def export_table_workbook(table_name: str, rows: list[dict[str, Any]]) -> BytesIO:
    wb=Workbook(); ws=wb.active; ws.title={"records":"SCHOOL_RECORDS","actions":"ACTION_PLANS","activity":"RECENT_ACTIVITY"}.get(table_name,"EXPORT")
    title={"records":"School Records Export","actions":"Action Plans Export","activity":"Recent Activity Export"}.get(table_name,"SRNHS Export")
    _title(ws,title,"Generated from current dashboard values.",12)
    if rows:
        headers=list(rows[0].keys()); _header(ws,5,headers)
        for r_index,row in enumerate(rows,6):
            for c_index,key in enumerate(headers,1): ws.cell(r_index,c_index,row.get(key,""))
        _format_table(ws,6,5+len(rows),1,len(headers))
        for col in range(1,len(headers)+1): ws.column_dimensions[get_column_letter(col)].width=min(max(18,len(str(headers[col-1]))+4),42)
    return _finalize(wb)


def import_workbook(file_stream) -> dict[str, Any]:
    workbook=load_workbook(file_stream,data_only=True)
    if "DATA_ENTRY" not in workbook.sheetnames: raise ValueError("The workbook must include a DATA_ENTRY sheet.")
    data_entry=workbook["DATA_ENTRY"]
    headers={str(data_entry.cell(5,col).value).strip():col for col in range(1,data_entry.max_column+1) if data_entry.cell(5,col).value}
    needed=["School Year","Level","Grade Level","Enrollment","Dropouts","Repeaters","Teachers"]
    if any(name not in headers for name in needed): raise ValueError("DATA_ENTRY is missing required school-record columns.")
    records=[]; years=set()
    for row in range(6,data_entry.max_row+1):
        year=str(data_entry.cell(row,headers["School Year"]).value or "").strip(); grade=str(data_entry.cell(row,headers["Grade Level"]).value or "").strip()
        if not year or not grade or data_entry.cell(row,headers["Enrollment"]).value in (None,""): continue
        item={"school_year":year,"level":str(data_entry.cell(row,headers["Level"]).value or "").strip(),"grade_level":grade,"enrollment":int(data_entry.cell(row,headers["Enrollment"]).value or 0),"dropouts":int(data_entry.cell(row,headers["Dropouts"]).value or 0),"repeaters":int(data_entry.cell(row,headers["Repeaters"]).value or 0),"teachers":int(data_entry.cell(row,headers["Teachers"]).value or 0)}
        records.append(item); years.add(year)
    resources=[]
    if "RESOURCES" in workbook.sheetnames:
        ws=workbook["RESOURCES"]
        for row in range(6,ws.max_row+1):
            year=str(ws.cell(row,1).value or "").strip()
            if re.fullmatch(r"\\d{4}-\\d{4}", year) and ws.cell(row,2).value not in (None,""):
                resources.append({"school_year":year,"jhs_classrooms":int(ws.cell(row,2).value or 0),"shs_classrooms":int(ws.cell(row,3).value or 0)})
    cohorts=[]
    if "COHORT_TRANSITION" in workbook.sheetnames:
        ws=workbook["COHORT_TRANSITION"]
        for row in range(6,ws.max_row+1):
            year=str(ws.cell(row,1).value or "").strip()
            if year and ws.cell(row,3).value not in (None,""): cohorts.append({"school_year":year,"baseline_year":str(ws.cell(row,2).value or ""),"grade7_baseline":int(ws.cell(row,3).value or 0),"grade12_current":int(ws.cell(row,4).value or 0)})
    actions=[]
    if "ACTION_PLANS" in workbook.sheetnames:
        ws=workbook["ACTION_PLANS"]; action_headers={str(ws.cell(5,col).value or "").strip():col for col in range(1,ws.max_column+1)}
        mapping={"School Year":"school_year","Analysis Type":"analysis_type","Focus Area":"focus_area","Observed Pattern":"observed_pattern","Data Basis":"data_basis","Suggested Action":"suggested_action","Responsible Group":"responsible_group","Target Indicator":"target_indicator","Baseline Value":"baseline_value","Target Value":"target_value","Monitoring Period":"monitoring_period","Current Result":"current_result","Status":"status","Progress Notes":"progress_notes","Additional Notes":"notes"}
        for row in range(6,ws.max_row+1):
            if not ws.cell(row, action_headers.get("Focus Area", 0) or 1).value: continue
            actions.append({dest:str(ws.cell(row,action_headers[src]).value or "") for src,dest in mapping.items() if src in action_headers})
    return {"records":records,"resources":resources,"cohorts":cohorts,"actions":actions,"years":sorted(years)}
